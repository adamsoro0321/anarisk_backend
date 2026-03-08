import pandas as pd
import logging
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows 
import xlsxwriter
import os
import networkx as nx
import matplotlib.pyplot as plt 
from matplotlib import animation
from datetime import date
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s - %(name)s - %(levelname)s - %(message)s",
    handlers=[
        logging.StreamHandler(),
        #logging.FileHandler(log_file_path, mode="a"),
    ],
)
class ProgrammeGenerator:
    OUTPUT_DIR = "../output"
    DATA_DIR = "../data"
    DOCS_DIR = "../docs"
    def __init__(self,
                quantume:pd.DataFrame ,
                ifu_data:pd.DataFrame,
                risk_data:pd.DataFrame,
                data_dgd:pd.DataFrame):
        self.logger = logging.getLogger(__name__)
        self.quantume = quantume
        self.ifu_data = ifu_data
        self.risk_data = risk_data
        self.data_dgd = data_dgd 
        self.load_reference_data()

    def load_reference_data(self):
        """Charger tous les fichiers de référence utilisés par les scripts R"""
        self.logger.info("Chargement des données de référence...")

        try:
            # 1. Liste des non-éligibles (NON_EIGIBLE.xlsx)
            reference_files = {
            "meta_donnees": f"{self.DOCS_DIR}/META_DONNEES.xlsx",
            "fiche_pistes": f"{self.DOCS_DIR}/FICHE_PISTES_INVESTIGATION.xlsx",
            "nomenclature_sh": f"{self.DOCS_DIR}/NomenclatureSH.xlsx",
            "dcf_activites": f"{self.DOCS_DIR}/DCF_ACTIVITES.xlsx",
            "suivi": f"{self.DOCS_DIR}/SUIVI.xlsx"
            }

            for attr_name, file_path in reference_files.items():
                if os.path.exists(file_path):
                    setattr(self, attr_name, pd.read_excel(file_path))
                    self.logger.info(f"{file_path} chargé")
                else:
                    setattr(self, attr_name, pd.DataFrame())
                    self.logger.warning(f"{file_path} non trouvé")

        except Exception as e:
            self.logger.error(
                f"Erreur lors du chargement des données de référence: {e}"
            )

    def process(self):
        # Placeholder for processing logic
        data = self.quantume.merge(self.risk_data,
                                on='NUM_IFU',
                                how='left').merge(self.ifu_data,
                                            on='NUM_IFU',
                                            how='inner')
        data["SOUS_UR"] = data["STRUCTURES_y"].str.replace(r'[ /]', '_', regex=True)
        data["UR"] = data["STRUCTURES_x"].str.replace(r'[ /]', '_', regex=True)
        data["BRIGADES"] = data["BRIGADES"].str.replace(r'[ /]', '_', regex=True)
        data["contribuables"] = data.apply(lambda row:f"{row['NUM_IFU']}_{row['SOUS_UR']}_{row['BRIGADES']}", axis=1)
        data["contribuables"] = data["contribuables"].str.replace(r'[^a-zA-Z0-9]', '_', regex=True)
        #traitement des dates
        data["DATE_DERNIERE_AVIS"] = pd.to_datetime(data["DATE_DERNIERE_AVIS"], errors='coerce')

        return data

    def generate(self,programme_root:str="../programmes"):
        data = self.process()
        for index, row in data.iterrows():
            # Placeholder for programme generation logic
            num_ifu = row['NUM_IFU']
            contribuable = row['contribuables']
            UR = row['UR']
            SUR = row['SOUS_UR']
            BV = row['BRIGADES']
            year1=2023
            year2=2024
            year3=2025
            entreprise_data = data[data['NUM_IFU'] == num_ifu] 

            wb = Workbook()

            # Créer la feuille Fiche_Contribuables
            ws_contribuables = wb.create_sheet("Fiche_Contribuables")
            ws_analyse_risk = wb.create_sheet("Fiche_Analyse_Risque")
            # Écrire le nom du contribuable en A1
            ws_contribuables.cell(1, 1, contribuable)
        
            # Définir les largeurs de colonnes (cols 2:4 = B, C, D)
            ws_contribuables.column_dimensions['B'].width = 40
            ws_contribuables.column_dimensions['C'].width = 40
            ws_contribuables.column_dimensions['D'].width = 40




            colonnes =data.columns.tolist()
            valeur_y1 = entreprise_data[entreprise_data["ANNEE_FISCAL"]==year1].reset_index(drop=True)
            valeur_y2 = entreprise_data[entreprise_data["ANNEE_FISCAL"]==year2].reset_index(drop=True)
            valeur_y3 = entreprise_data[entreprise_data["ANNEE_FISCAL"]==year3].reset_index(drop=True)

            #remplir frame fiche contribuables
            for i ,col in enumerate(colonnes):
                ws_contribuables.cell(i+1,2, col)
                if not valeur_y1.empty:
                    y = valeur_y1.iloc[0][col]
                    ws_contribuables.cell(i+1,3, y)
                if not valeur_y2.empty:
                    y = valeur_y2.iloc[0][col]
                    ws_contribuables.cell(i+1,4, y)
                if not valeur_y3.empty:
                    y = valeur_y3.iloc[0][col]
                    ws_contribuables.cell(i+1,5, y)

            #2.remplir frame analyse risque
            ws_analyse_risk.cell(1, 1, contribuable)
            for i, col in enumerate(colonnes):
                x =colonnes[i]
                ws_analyse_risk.cell( i+1,2, col)
                if not valeur_y1.empty:
                    y = valeur_y1.iloc[0][col]
                    ws_analyse_risk.cell( i+1,3, y)
            z="Confirmation  (1= total,2=partielle,3=pas de confirmation)"
            ws_analyse_risk.cell(1,5, z)
            if self.suivi is not None:
                ws_analyse_risk.cell(2,7, str(self.suivi.to_dict()))


            programme_path = f"{programme_root}/programme_{date.today()}/{UR}/{SUR}/{BV}"
            #cree le repertoire si n'existe pas

            os.makedirs(programme_path, exist_ok=True)
            wb.save(f"{programme_path}/programme_{contribuable}.xlsx")

            self.logger.info(f"Generated programme for IFU {num_ifu} at {programme_path}")
        return f"Generated programme: "